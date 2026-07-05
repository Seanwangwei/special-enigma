"""Generate the UAT test cases Excel file at docs/UAT_Test_Cases.xlsx."""

from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

TEST_CASES = [
    ("TC-01", "Launch & Empty State", [
        ("1", "Launch python3 app.py", 'Window appears titled "Exam Result Email Automation", ~680×680px'),
        ("2", "Observe the main window", 'Four sections visible: "📄 Email Template (.docx)", "📊 Excel File", "📎 Attachments", "⚡ Actions"'),
        ("3", "Check action buttons", '"▶ Preview Emails", "✉ Send Test Email", "📨 Send All Emails" are all greyed out / disabled'),
        ("4", "Check progress & log", '"📈 Progress" and "📋 Log" sections hidden'),
        ("5", "Check status bar", '"Ready" with green styling'),
    ]),
    ("TC-02", "Window Resize & Geometry Persistence", [
        ("1", "Resize the window smaller", "Cannot go below ~480×400px"),
        ("2", "Resize to a custom size, then close & relaunch the app", "Window reopens at the last size & position"),
    ]),
    ("TC-03", "DOCX Template Upload", [
        ("1", 'Click "Browse…" in the Template section', "File dialog opens, filtered to .docx files"),
        ("2", "Select a .docx with {{variables}}", "File path appears in the read-only text field"),
        ("3", "Observe variable badges", "Blue info badges appear below the file path, one per {{variable}} detected"),
        ("4", "Observe variable count label", 'Shows "N variable(s) detected" (and "Special: module_table" if applicable)'),
        ("5", "Observe log output", "Log panel appears (if hidden) with: Loaded template: <name> | Subject: <subject> | N variable(s), M special"),
        ("6", "Observe status bar", 'Updates to "Template loaded. Select an Excel file to continue."'),
    ]),
    ("TC-04", "DOCX Template — Subject Prompt", [
        ("1", "Upload a .docx with no companion .yaml", 'A dialog pops up asking "Enter the email subject line for this template:"'),
        ("2", "Enter a subject and click OK", "Template loads with that subject"),
        ("3", "Repeat with a .docx that has subject: in companion .yaml", "No subject prompt — uses the YAML subject directly"),
    ]),
    ("TC-05", "Clear Template", [
        ("1", 'After TC-03, click "Clear" in the Template section', 'Badges and variable count label disappear; path field cleared; status bar returns to "Ready"'),
        ("2", "Observe log", '"Uploaded template cleared. Using bundled HTML templates."'),
    ]),
    ("TC-06", "Excel File Load — With Template (Match)", [
        ("1", "Upload a .docx template (TC-03 again)", "Template loaded"),
        ("2", 'Click "Browse…" in Excel section, pick an .xlsx whose columns match the template variables', "File path appears; progress bar & log panel become visible"),
        ("3", "Observe progress bar", 'Updates as previews generate (e.g. "15 / 15")'),
        ("4", "Observe log", 'Loaded N students from <path>, then per-student "Generating preview…", then "Email previews generated."'),
        ("5", "Check action buttons", "All three enabled: Preview, Send Test, Send All"),
        ("6", "Check status bar", '"Excel file loaded successfully."'),
    ]),
    ("TC-07", "Excel File Load — Column Mismatch", [
        ("1", "Upload a template with variables that don't match the Excel columns", 'A "Column Mismatch" warning dialog appears'),
        ("2", "Observe dialog content", 'Lists "Missing from Excel" variables and/or "Unused Excel columns"'),
        ("3", 'Click "No" (default)', 'Dialog closes, Excel is NOT loaded, log says "Excel load cancelled due to column mismatch."'),
        ("4", 'Repeat, this time click "Yes"', "Excel loads despite mismatch; missing variables will render as empty in emails"),
    ]),
    ("TC-08", "Attachments", [
        ("1", 'Click "Browse…" in Attachments section', "Multi-file dialog opens"),
        ("2", "Select 2 files", "Paths appear separated by ; "),
        ("3", 'Click "Browse…" again, add a 3rd file', "All three paths shown"),
        ("4", 'Click "Clear"', 'Path field clears, log says "Attachments cleared."'),
    ]),
    ("TC-09", "Preview Emails Dialog", [
        ("1", 'With students loaded (TC-06), click "▶ Preview Emails"', "Preview dialog opens centered on main window"),
        ("2", "Observe layout", "Left pane: student list; Right pane: rendered HTML email"),
        ("3", "Click a student in the left list", "Right pane updates to show that student's email (To, Subject + HTML body)"),
        ("4", "Verify different students", "Each shows personalized content (name, modules, etc.)"),
        ("5", "Close the dialog", "Returns to main window"),
    ]),
    ("TC-10", "Send Test Email — Single Student", [
        ("1", 'Click "✉ Send Test Email"', 'Dialog: "Select Student" dropdown with all loaded students'),
        ("2", "Pick a student, click OK", "Delivery Method dialog opens (Outlook / SMTP radio cards)"),
        ("3", "Select method, click Continue", "Delivery Mode dialog opens (Preview / Draft / Send radio cards)"),
        ("4", "Select mode, click Continue", "Email processed for that single student"),
        ("5", "Observe log", 'Test email Sent/Draft Created/Preview Generated for <name>'),
        ("6", "Observe status bar", '"Test email Sent" (or relevant status)'),
    ]),
    ("TC-11", "Send All — Full Wizard (Happy Path)", [
        ("1", 'Click "📨 Send All Emails"', "If any validation errors → Validation dialog first (see TC-12); otherwise goes to Delivery Method"),
        ("2", "Delivery Method dialog", "Two radio cards: Outlook (with account dropdown) and SMTP; Step indicator shows Step 1 active"),
        ("3", "Select Outlook, click Continue", "Delivery Mode dialog opens"),
        ("4", "Delivery Mode dialog", "Three radio cards: 🔍 Preview Only, 📝 Create Drafts (pre-selected), 📨 Send Immediately; Step indicator shows Step 2 active"),
        ("5", 'Select "Create Drafts", click Continue', "Confirmation dialog opens"),
        ("6", "Confirmation dialog", "Shows student count, Pass/Fail/EC breakdown, Delivery Method + Mode; Step indicator shows Step 3 active"),
        ("7", 'Click "📨 Process N Emails"', 'Emails processed in background; progress bar updates in real-time; log shows "Sending N of M: <name>"'),
        ("8", "On completion", "Results dialog appears with tabs (All Records + Errors); audit log saved to logs/"),
    ]),
    ("TC-12", "Send All — Validation Errors", [
        ("1", "Load an Excel with missing emails or duplicate Student IDs", "(requires a crafted test file)"),
        ("2", 'Click "📨 Send All Emails"', "Validation Results dialog appears: lists missing emails and/or duplicate IDs with red/amber styling"),
        ("3", 'Click "Go Back & Fix"', "Dialog closes, send cancelled, log confirms cancellation"),
        ("4", 'Click "Continue Anyway"', "Proceeds to Delivery Method dialog as normal"),
    ]),
    ("TC-13", "Cancel at Each Wizard Step", [
        ("1", "At Delivery Method dialog → click Back or close (X)", 'Send cancelled, log: "Send cancelled at delivery method selection."'),
        ("2", "At Delivery Mode dialog → click Back or close (X)", 'Send cancelled, log: "Send cancelled at delivery mode selection."'),
        ("3", "At Confirmation dialog → click Back or close (X)", 'Send cancelled, log: "Send cancelled at confirmation."'),
    ]),
    ("TC-14", "Keyboard Shortcuts", [
        ("1", "Press Ctrl+O", "Excel file browse dialog opens"),
        ("2", "Press Ctrl+P (with students loaded)", "Preview dialog opens"),
        ("3", "Press Ctrl+P (no students loaded)", 'Error message: "Please load an Excel file first."'),
    ]),
    ("TC-15", "Clear Excel — State Reset", [
        ("1", 'After loading students, click "Clear" in the Excel section', 'Path clears; action buttons disable; progress & log hide; log says "Excel file cleared."; status bar: "Ready"'),
    ]),
    ("TC-16", "Template Change Clears Data", [
        ("1", "Load a template + Excel (TC-06)", "Students loaded, previews generated"),
        ("2", "Upload a different .docx template", 'Log says "Template changed — previously loaded student data has been cleared."'),
        ("3", "Verify", "Student data, previews, progress all reset; buttons disabled"),
    ]),
    ("TC-17", "Send Test Email — Cancel at Each Step", [
        ("1", 'Click "✉ Send Test Email", pick student, click OK', "Delivery Method opens"),
        ("2", "Close the dialog (X)", "Returns to main window, nothing sent, no errors"),
        ("3", "Repeat through Delivery Mode dialog, close it", "Same — clean cancellation"),
    ]),
    ("TC-18", "Design Conformance — Quick Checks", [
        ("1", 'Buttons: "Browse…"', "Secondary style (white bg, border)"),
        ("2", 'Buttons: "Clear"', "Ghost style (transparent, grey text)"),
        ("3", 'Buttons: "▶ Preview Emails"', "Primary style (blue bg, white text)"),
        ("4", "Section headers", 'QGroupBox titles: "📄 Email Template (.docx)", "📊 Excel File", "📎 Attachments", "⚡ Actions"'),
        ("5", "Variable badges", "Small blue pills with variable names"),
        ("6", "Radio cards (delivery dialogs)", "Border on all sides, blue highlight when selected"),
        ("7", "Log panel", "Dark background, monospace light text"),
        ("8", "Progress bar", 'Blue fill, shows "N / M" label below'),
        ("9", "All dialogs", "Centered on parent window, not fixed position"),
    ]),
    ("TC-19", "Edge Cases — Guard Actions Without Data", [
        ("1", 'Click "▶ Preview Emails" with no Excel loaded', 'Error dialog: "Please load an Excel file first."'),
        ("2", 'Click "✉ Send Test Email" with no Excel loaded', 'Error dialog: "Please load an Excel file first."'),
        ("3", 'Click "📨 Send All Emails" with no Excel loaded', 'Error dialog: "Please load an Excel file first."'),
    ]),
]

# ----- Styles -----
HEADER_FILL = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TC_ROW_FILL = PatternFill(start_color="F0F5FF", end_color="F0F5FF", fill_type="solid")
TC_FONT = Font(name="Calibri", size=11, bold=True, color="1E3A6E")
STEP_FONT = Font(name="Calibri", size=10, color="374151")
BODY_FONT = Font(name="Calibri", size=10, color="374151")
ACTION_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def build(workbook: Workbook) -> None:
    ws = workbook.active
    ws.title = "UAT Test Cases"

    # Column widths
    ws.column_dimensions["A"].width = 10   # TC ID
    ws.column_dimensions["B"].width = 12   # Step
    ws.column_dimensions["C"].width = 52   # Action
    ws.column_dimensions["D"].width = 58   # Expected Behaviour
    ws.column_dimensions["E"].width = 16   # Result
    ws.column_dimensions["F"].width = 30   # Notes

    # Freeze panes (header + first data row visible)
    ws.freeze_panes = "A2"

    # ---- Header row ----
    headers = ["Test Case", "Step", "Action", "Expected Behaviour", "Result", "Notes"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 28

    # ---- Data rows ----
    row = 2
    for tc_id, tc_title, steps in TEST_CASES:
        for step_num, action, expected in steps:
            ws.cell(row=row, column=1, value=tc_id).font = TC_FONT
            ws.cell(row=row, column=2, value=step_num).font = STEP_FONT
            ws.cell(row=row, column=3, value=action).font = BODY_FONT
            ws.cell(row=row, column=4, value=expected).font = BODY_FONT
            # Result & Notes left blank for manual entry
            ws.cell(row=row, column=5, value="").font = BODY_FONT
            ws.cell(row=row, column=6, value="").font = BODY_FONT

            for col_idx in range(1, 7):
                cell = ws.cell(row=row, column=col_idx)
                cell.border = THIN_BORDER
                cell.alignment = WRAP
                # Alternate TC background
                tc_index = [t[0] for t in TEST_CASES].index(tc_id)
                if tc_index % 2 == 0:
                    cell.fill = ACTION_FILL

            row += 1

    # ---- Data validation: Result dropdown ----
    dv = DataValidation(
        type="list",
        formula1='"✅ Pass,❌ Fail,⚠️ Minor Issue,🚧 Blocked,🔳 Not Tested"',
        allow_blank=True,
    )
    dv.error = "Please select a valid result."
    dv.errorTitle = "Invalid Result"
    ws.add_data_validation(dv)
    dv.add(f"E2:E{row - 1}")

    # ---- Auto-filter ----
    ws.auto_filter.ref = f"A1:F{row - 1}"

    # ---- Metadata sheet ----
    ws2 = workbook.create_sheet("Info")
    info_data = [
        ("Project", "Exam Result Email Automation"),
        ("Version", "v0.6.0"),
        ("Test Date", "July 5, 2026"),
        ("Tester", "Sean Wang"),
        ("Total Test Cases", str(len(TEST_CASES))),
        ("Total Steps", str(sum(len(steps) for _, _, steps in TEST_CASES))),
        ("Result Options", "✅ Pass | ❌ Fail | ⚠️ Minor Issue | 🚧 Blocked | 🔳 Not Tested"),
    ]
    for i, (label, value) in enumerate(info_data, start=1):
        ws2.cell(row=i, column=1, value=label).font = Font(name="Calibri", size=11, bold=True)
        ws2.cell(row=i, column=2, value=value).font = Font(name="Calibri", size=11)
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 60


if __name__ == "__main__":
    wb = Workbook()
    build(wb)
    output = Path(__file__).resolve().parent.parent / "docs" / "UAT_Test_Cases.xlsx"
    wb.save(output)
    print(f"✅ Saved {len(TEST_CASES)} test cases, {sum(len(s) for _, _, s in TEST_CASES)} steps → {output}")
